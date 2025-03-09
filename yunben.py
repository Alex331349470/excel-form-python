import pandas as pd
import os

# 设置文件路径
excel_dir = os.path.join(os.path.dirname(__file__), 'excel_files')
source_file = os.path.join(excel_dir, '云本农发行.xlsx')
target_file = os.path.join(excel_dir, '资金日报汇总表.xlsx')

print(source_file)
print(f"reading {source_file}")
source_df = pd.read_excel(source_file, header=4)

column_names = source_df.columns
print(column_names.tolist())

if "账户余额" in column_names:
    first_balance = source_df["账户余额"].iloc[0]
    print(f"第一行账户余额: {first_balance}")
    last_balance = source_df["账户余额"].iloc[-1]
    print(f"最后一行账户余额: {last_balance}")
else:
    print("账户余额列不存在,可用的字段有:", column_names.tolist())

# 将first_balance和last_balance写入target_file的D3和G3单元格
if "账户余额" in column_names:
    try:
        from openpyxl import load_workbook
        
        # 加载目标工作簿
        print(f"正在打开目标文件: {target_file}")
        workbook = load_workbook(target_file)
        
        # 选择"云本"工作表
        if "云本" in workbook.sheetnames:
            sheet = workbook["云本"]
            
            # 写入账户余额数据
            sheet['D3'] = first_balance
            sheet['G3'] = last_balance
            
            # 保存工作簿
            workbook.save(target_file)
            print(f"成功将数据写入{target_file}的'云本'工作表：")
            print(f"单元格D3: {first_balance}")
            print(f"单元格G3: {last_balance}")
        else:
            print(f"错误：目标文件中不存在'云本'工作表！")
            print(f"可用的工作表: {workbook.sheetnames}")
    except Exception as e:
        print(f"写入Excel文件时出错: {e}")
else:
    print("由于账户余额列不存在，无法写入目标文件")

# 将source_file数据填入target_file中的农业发展银行相关字段
if "贷方发生额" in column_names:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        
        # 获取需要的数据列
        required_columns = ["贷方发生额", "交易对手名称", "摘要", "用途"]
        data_dict = {}
        
        # 检查每个列是否存在并获取数据
        for col_name in required_columns:
            if col_name in column_names:
                # 过滤掉NaN值
                data_dict[col_name] = source_df[col_name].dropna().tolist()
                print(f"获取到{len(data_dict[col_name])}个{col_name}数据")
            else:
                data_dict[col_name] = []
                print(f"警告: 源文件中不存在'{col_name}'列")
        
        # 确定要处理的行数 (使用贷方发生额的行数作为基准)
        if not data_dict["贷方发生额"]:
            print("没有贷方发生额数据，无法继续处理")
            raise ValueError("没有贷方发生额数据")
            
        num_rows = len(data_dict["贷方发生额"])
        
        # 加载目标工作簿
        workbook = load_workbook(target_file)
        
        # 确认云本工作表存在
        if "云本" in workbook.sheetnames:
            sheet = workbook["云本"]
            
            # 查找"收款"和"农业发展银行"所在区域
            receipt_row = None  # "收款"所在行
            agri_bank_rows = []  # "农业发展银行"所在的行
            target_cols = {
                "序号": None,
                "金额": None,
                "往来单位": None, 
                "摘要": None,
                "备注": None
            }
            
            # 查找"收款"行
            for row in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value and "收款" == str(cell_value).strip():
                    receipt_row = row
                    break
            
            if receipt_row:
                # 查找"农业发展银行"行
                for row in range(receipt_row, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row, column=2).value
                    if cell_value and "农业发展银行" in str(cell_value):
                        agri_bank_rows.append(row)
                    elif agri_bank_rows and cell_value and cell_value.strip():  # 如果已经找到农发行行且当前B列有其他内容
                        break  # 说明农发行部分结束了
                
                if agri_bank_rows:
                    # 检查第一行是否包含字段名
                    first_bank_row = agri_bank_rows[0]
                    
                    # 查找各个目标列
                    # 1. 先在第一行中查找
                    for col in range(1, sheet.max_column + 1):
                        cell_value = sheet.cell(row=first_bank_row, column=col).value
                        if not cell_value:
                            continue
                            
                        cell_str = str(cell_value).strip()
                        for field in target_cols.keys():
                            if field in cell_str and target_cols[field] is None:
                                target_cols[field] = col
                    
                    # 2. 检查表头行
                    if any(col is None for col in target_cols.values()):
                        header_rows = list(range(max(1, receipt_row - 3), receipt_row))
                        for header_row in header_rows:
                            for col in range(1, sheet.max_column + 1):
                                cell_value = sheet.cell(row=header_row, column=col).value
                                if not cell_value:
                                    continue
                                    
                                cell_str = str(cell_value).strip()
                                for field in target_cols.keys():
                                    if field in cell_str and target_cols[field] is None:
                                        target_cols[field] = col
                    
                    # 默认值处理
                    if target_cols["序号"] is None:
                        # 通常序号在最左边几列
                        for col in range(1, 5):
                            if col not in target_cols.values():
                                target_cols["序号"] = col
                                break
                    
                    if target_cols["金额"] is None:
                        target_cols["金额"] = 3  # 默认C列
                        
                    if target_cols["往来单位"] is None:
                        target_cols["往来单位"] = 4  # 默认D列
                        
                    if target_cols["摘要"] is None:
                        target_cols["摘要"] = 5  # 默认E列
                        
                    if target_cols["备注"] is None:
                        target_cols["备注"] = 6  # 默认F列
                    
                    print(f"找到农业发展银行行: {agri_bank_rows}")
                    print(f"目标列: 序号={get_column_letter(target_cols['序号'])}, "
                          f"金额={get_column_letter(target_cols['金额'])}, "
                          f"往来单位={get_column_letter(target_cols['往来单位'])}, "
                          f"摘要={get_column_letter(target_cols['摘要'])}, "
                          f"备注={get_column_letter(target_cols['备注'])}")
                    
                    # 现有的农发行行数
                    existing_rows = len(agri_bank_rows)
                    
                    # 考虑第一行可能是复合字段名，从第二行开始填充
                    start_row = 0  # 默认从第一行开始填充
                    first_row_is_header = False
                    
                    # 检查第一行是否是表头
                    for field, col in target_cols.items():
                        first_row_cell = sheet.cell(row=first_bank_row, column=col).value
                        if first_row_cell and field in str(first_row_cell):
                            first_row_is_header = True
                            break
                    
                    if first_row_is_header:
                        start_row = 1  # 从索引1开始（第二行）
                        if len(agri_bank_rows) <= 1:
                            # 仅有一行且是字段行，需要新增数据行
                            last_row = agri_bank_rows[0]
                            sheet.insert_rows(last_row + 1)
                            agri_bank_rows.append(last_row + 1)
                            sheet.cell(row=last_row + 1, column=2).value = "农业发展银行"
                            existing_rows += 1
                    
                    # 填充现有行
                    rows_to_fill = min(existing_rows - start_row, num_rows)
                    for i in range(rows_to_fill):
                        row_idx = agri_bank_rows[start_row + i]
                        
                        # 序号从1开始递增
                        sheet.cell(row=row_idx, column=target_cols["序号"]).value = i + 1
                        
                        # 金额
                        if i < len(data_dict["贷方发生额"]):
                            sheet.cell(row=row_idx, column=target_cols["金额"]).value = data_dict["贷方发生额"][i]
                        
                        # 往来单位
                        if "交易对手名称" in data_dict and i < len(data_dict["交易对手名称"]):
                            sheet.cell(row=row_idx, column=target_cols["往来单位"]).value = data_dict["交易对手名称"][i]
                        
                        # 摘要
                        if "摘要" in data_dict and i < len(data_dict["摘要"]):
                            sheet.cell(row=row_idx, column=target_cols["摘要"]).value = data_dict["摘要"][i]
                        
                        # 备注
                        if "用途" in data_dict and i < len(data_dict["用途"]):
                            sheet.cell(row=row_idx, column=target_cols["备注"]).value = data_dict["用途"][i]
                    
                    # 如果有更多的数据需要填充，添加新行
                    if num_rows > (existing_rows - start_row):
                        last_row = agri_bank_rows[-1]
                        for i in range(existing_rows - start_row, num_rows):
                            # 插入新行
                            sheet.insert_rows(last_row + 1)
                            
                            # 更新后续行的索引
                            agri_bank_rows = [r if r <= last_row else r + 1 for r in agri_bank_rows]
                            last_row += 1
                            agri_bank_rows.append(last_row)
                            
                            # 在B列填入"农业发展银行"
                            sheet.cell(row=last_row, column=2).value = "农业发展银行"
                            
                            # 序号递增
                            sheet.cell(row=last_row, column=target_cols["序号"]).value = i + 1
                            
                            # 金额
                            if i < len(data_dict["贷方发生额"]):
                                sheet.cell(row=last_row, column=target_cols["金额"]).value = data_dict["贷方发生额"][i]
                            
                            # 往来单位
                            if "交易对手名称" in data_dict and i < len(data_dict["交易对手名称"]):
                                sheet.cell(row=last_row, column=target_cols["往来单位"]).value = data_dict["交易对手名称"][i]
                            
                            # 摘要
                            if "摘要" in data_dict and i < len(data_dict["摘要"]):
                                sheet.cell(row=last_row, column=target_cols["摘要"]).value = data_dict["摘要"][i]
                            
                            # 备注
                            if "用途" in data_dict and i < len(data_dict["用途"]):
                                sheet.cell(row=last_row, column=target_cols["备注"]).value = data_dict["用途"][i]
                    
                    # 保存工作簿
                    workbook.save(target_file)
                    print(f"成功将{num_rows}条数据写入{target_file}的'云本'工作表")
                else:
                    print("未找到'农业发展银行'行")
            else:
                print("未找到'收款'行")
        else:
            print(f"目标文件中不存在'云本'工作表")
    except Exception as e:
        import traceback
        print(f"写入数据时出错: {e}")
        traceback.print_exc()
else:
    print("贷方发生额列不存在,可用的字段有:", column_names.tolist())
